import io
from flask import Flask, abort, render_template, request, send_file, session, Response
from PyPDF2 import PdfFileReader
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
import Form

app = Flask(__name__)
app.secret_key = '''eb ef 28 c1 fd 42 00 3b a8 7e 91 d9 eb 0f 19 88 
c3 7a 33 04 6f a7 0c 9c db 52 7a 3f bf 66 2f 3b 
42 82 48 93 98 3e 22 5b ca 91 7b d9 5c e1 b6 d6 
aa 87 92 53 f7 46 f9 69 7c 72 4f 2c 65 31 22 ab 
45 76 2c 43 68 d3 93 d4 60 c1 6a ab e7 b3 af 1e 
58 7c 78 7d 07 92 64 4c a4 f7 08 16 4f 2f 22 53 
2b 4e c0 08'''


class Form:

    X = "X"

    def __init__(self, title, df):

        self.title = title

        if df['Check Box20'] == self.X:
            self.Unterkunft = 'Hotel Premium'
        elif df['Check Box21'] == self.X:
            self.Unterkunft = 'Hotel Mittelklasse' 
        elif df['Check Box32'] == self.X:
            self.Unterkunft = 'Hotel/Hostel' 
        elif df['Check Box19'] == self.X:
            self.Unterkunft = 'Hütte nicht DAV' 
        elif df['Check Box18'] == self.X:
            self.Unterkunft = 'DAV-Hütte' 
        elif df['Check Box17'] == self.X:
            self.Unterkunft = 'Camping' 
        elif df['Check Box16'] == self.X:
            self.Unterkunft = 'Biwak' 
        else:    
            self.Unterkunft = '' 

        if df['Ja'] == self.X:
            self.Besitz = 'Ja' 
        elif df['Nein'] == self.X:
            self.Besitz = 'Nein' 
        else:    
            self.Besitz = '' 

        if df['andere Räumlichkeiten zB Tagungsraum Eventlocation'] == self.X:
            self.Ort = 'Tagungsraum/Hotel' 
        elif df['KletterBoulderhalle'] == self.X:
            self.Ort = 'Kletterhalle/Boulderhalle' 
        elif df['Außenanlage zB Außenkletterturm'] == self.X:
            self.Ort = 'Außenanlage' 
        elif df['Naturraumim freien'] == self.X:
            self.Ort = 'Naturraum' 
        else:    
            self.Ort = '' 

        self.data_dict = {
            'DAV-Sektion': '',
            'Veranstaltungs_Gruppenname': df['VeranstaltungsGruppenname'],
            'Referat_Geschäftsbereich_Abteilung': df['ReferatGeschäftsbereichAbteilung'],
            'Sportart': df['Sportart'],
            'Anzahl_gleicher_Aktivitäten': 1,
            'Datum_der_Aktivität': df['Datum'],
            'Dauer_in_Tagen': df['Dauer in Tagen 05 Schritte'],
            'Anzahl_Teilnehmer': df['Anzahl Teilnehmerinnen inkl Leitung'],
            'Veranstaltungsort': df['Veranstaltungsort'],
            'Land': df['Land'],
            'Veranstaltungstyp': '',
            'Anfahrt_Diesel_Benzin_PKW': df['Text21'],
            'Anfahrt_Elektro_PKW': df['Text22'],
            'Anfahrt_Van_Transporter': df['Text23'],
            'Anfahrt_Sektionsbus': df['Text24'],
            'Anfahrt_Mitfahrer_innen': df['Text36'],
            'Anfahrt_Fahrrad_zu_Fuß': df['Text14'],
            'Anfahrt_ÖPNV': df['Text20'],
            'Anfahrt_Fernverkehr_Zug': df['Text19'],
            'Anfahrt_Reisebus': df['Text18'],
            'Flug_Anzahl_Personen ': df['Anzahl Personen'],
            'Flug_Anzahl_Flugsegmente': 1,
            'Startflughafen': df['Start Flughafen'],
            'Zielflughafen': df['Ziel Flughafen'],
            'Mobilität_Diesel_Benzin_PKW': float(df['Text47']) * float(df['Text22']),
            'Mobilität_Elektro_PKW': float(df['Text33']) * float(df['Text38']),
            'Mobilität_Van_Transporter': float(df['Text34']) * float(df['Text39']),
            'Mobilität_Sektionsbus': float(df['Text35']) * float(df['Text30']),
            'Mobilität_ÖPNV': df['Text31'],
            'Gondel_Lift': df['GondelLift Anzahl Fahrten pro Person'],
            'Anzahl_Übernachtung ': float(df['Text40']) + float(df['Text41']) + float(df['Text42']) + float(df['Text43']) + float(df['Text44']) + float(df['Text45']),
            'Art_der_Unterkunft': self.Unterkunft,
            'Anzahl_servierter_Mahlzeiten': df['Anzahl servierter Mahlzeiten pro Person'],
            'Prozent vegane Mahlzeiten': df['Anteilig'],
            'Prozent vegetarische Mahlzeiten': df['vegetarisch'],
            'Prozent Mahlzeiten mit Fleisch': df['mit Fleisch'],
            'Besitzverhältnis': self.Besitz,
            'Veranstaltungsort': self.Ort,
            'Größe_Tagungsraum': df['Größe in m²'],
        }

        self.row_dict = {
            5: '',
            6: df['VeranstaltungsGruppenname'],
            7: df['ReferatGeschäftsbereichAbteilung'],
            8: df['Sportart'],
            9: 1,
            10: df['Datum'],
            11: df['Dauer in Tagen 05 Schritte'],
            12: df['Anzahl Teilnehmerinnen inkl Leitung'],
            13: df['Veranstaltungsort'],
            14: df['Land'],
            15: '',
            17: df['Text21'],
            18: df['Text22'],
            19: df['Text23'],
            20: df['Text24'],
            21: df['Text36'],
            22: df['Text14'],
            23: df['Text20'],
            24: df['Text19'],
            25: df['Text18'],
            38: df['Anzahl Personen'],
            38: "",
            40: df['Start Flughafen'],
            41: df['Ziel Flughafen'],
            43: float(df['Text47']) * float(df['Text22']),
            44: float(df['Text33']) * float(df['Text38']),
            45: float(df['Text34']) * float(df['Text39']),
            46: float(df['Text35']) * float(df['Text30']),
            47: df['Text31'],
            48: df['GondelLift Anzahl Fahrten pro Person'],
            50: float(df['Text40']) + float(df['Text41']) + float(df['Text42']) + float(df['Text43']) + float(df['Text44']) + float(df['Text45']),
            51: self.Unterkunft,
            60: df['Anzahl servierter Mahlzeiten pro Person'],
            61: df['Anteilig'],
            62: df['vegetarisch'],
            63: df['mit Fleisch'],
            65: self.Besitz,
            66: self.Ort,
            67: df['Größe in m²'],
        }

    def data_df(self):
        return pd.DataFrame.from_dict(self.data_dict, orient='index', columns = [self.title]).replace(np.nan, "")
    def row_df(self):
        return pd.DataFrame.from_dict(self.row_dict, orient='index', columns = [self.title]).replace(np.nan, "")

@app.route('/')
def main():
    return render_template("index.html")
  
  
@app.route('/upload', methods=['POST'])
def upload():
    if request.method == 'POST':
  
        res_dict = {}
        # Get the list of files from webpage
        files = request.files.getlist("file")

        for i, file in enumerate(files):
            pdf_reader = PdfFileReader(file)
    
            fields = pdf_reader.get_fields()
            df = pd.DataFrame.from_dict(fields).T[['/FT', '/V']]
            df.loc[df['/FT'] == "/Btn", '/V'] = df.loc[df['/FT'] == "/Btn", '/V'].replace("/Off", np.nan).replace(['/On', '/Ja'], "X")
            res_dict["form_" + str(i) + "_" + file.name] = df

        res_df = pd.concat(res_dict)
        res_df.index.names = ["form", 'feld']
        out_df = res_df.reset_index().pivot(index='feld', columns='form', values='/V')
        data_list = []
        row_list = []
        for col in out_df.columns:
            data = Form(col, out_df.to_dict()[col])
            data_list.append(data.data_df())
            row_list.append(data.row_df())

        table_df = pd.concat(data_list, axis=1)
        row_df = pd.concat(row_list, axis=1)

        session["df"] = table_df.to_csv(index=True, header=True, sep=";")
        session['table'] = row_df.reset_index().to_dict('dict')

        return render_template("overview.html", result_table = table_df.to_html(classes=['responsive-table', 'striped']))

@app.route("/get_csv", methods=['GET', 'POST'])
def get_csv():
    
    if "df" in session:
        csv = session["df"]
    else:
        abort(404)

    
    
    # Create a string buffer
    buf_str = io.StringIO(csv)

    # Create a bytes buffer from the string buffer
    buf_byt = io.BytesIO(buf_str.read().encode("utf-8"))
    
    # Return the CSV data as an attachment
    return send_file(buf_byt,
                     mimetype="text/csv",
                     as_attachment=True,
                     attachment_filename="data.csv")


@app.route("/get_xlsx", methods=['GET', 'POST'])
def get_xlsx():
    
    if "table" not in session:
        abort(404)

    df = pd.DataFrame(session['table']).set_index('index')
    print(df)
    wb = load_workbook(filename = "Veranstaltungsaktivitaeten.xlsx")
    ws = wb.active
    col = 6
    for colname in df.columns:
        print(colname)
        for rowname in df[colname].index:
            print(rowname)
            _ = ws.cell(column=col, row=rowname, value=df.loc[rowname, colname])
        col += 1

    #wb.save(filename = 'out.xlsx')
    
    return Response(
        save_virtual_workbook(wb),
        headers={
            'Content-Disposition': 'attachment; filename=sheet.xlsx',
            'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    )

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8080, debug=True)